from datetime import date
from io import BytesIO

from django.test import TestCase
from django.urls import reverse

from .models import Author, Category, Book
from .services import import_books_from_excel, export_books_to_excel

from openpyxl import Workbook


class AuthorModelTest(TestCase):

    def test_author_creation(self):
        author = Author.objects.create(
            first_name='Джордж',
            last_name='Оруэлл',
            biography='Британский писатель'
        )

        self.assertEqual(author.first_name, 'Джордж')
        self.assertEqual(author.last_name, 'Оруэлл')

    def test_author_str(self):
        author = Author.objects.create(
            first_name='Джордж',
            last_name='Оруэлл'
        )

        self.assertEqual(
            str(author),
            'Джордж Оруэлл'
        )


class CategoryModelTest(TestCase):

    def test_category_creation(self):
        category = Category.objects.create(
            name='Роман',
            description='Художественная литература'
        )

        self.assertEqual(category.name, 'Роман')

    def test_category_str(self):
        category = Category.objects.create(
            name='Роман'
        )

        self.assertEqual(
            str(category),
            'Роман'
        )


class BookModelTest(TestCase):

    def setUp(self):
        self.author = Author.objects.create(
            first_name='Джордж',
            last_name='Оруэлл'
        )

        self.category = Category.objects.create(
            name='Роман'
        )

    def test_book_creation(self):
        book = Book.objects.create(
            title='1984',
            isbn='9780451524935',
            description='Роман-антиутопия',
            price=340,
            stock_quantity=10,
            publication_date=date(1949, 6, 8)
        )

        book.authors.add(self.author)
        book.categories.add(self.category)

        self.assertEqual(book.title, '1984')
        self.assertEqual(book.isbn, '9780451524935')

    def test_book_str(self):
        book = Book.objects.create(
            title='1984',
            isbn='9780451524935',
            description='Роман-антиутопия',
            price=340,
            stock_quantity=10,
            publication_date=date(1949, 6, 8)
        )

        self.assertEqual(
            str(book),
            '1984'
        )

    def test_book_relations(self):
        book = Book.objects.create(
            title='1984',
            isbn='9780451524935',
            description='Роман-антиутопия',
            price=340,
            stock_quantity=10,
            publication_date=date(1949, 6, 8)
        )

        book.authors.add(self.author)
        book.categories.add(self.category)

        self.assertEqual(
            book.authors.count(),
            1
        )

        self.assertEqual(
            book.categories.count(),
            1
        )


class CatalogViewsTest(TestCase):

    def setUp(self):
        self.author = Author.objects.create(
            first_name='Лев',
            last_name='Толстой'
        )

        self.category = Category.objects.create(
            name='Роман'
        )

        self.book = Book.objects.create(
            title='Война и мир',
            isbn='9785170901234',
            description='Роман',
            price=400,
            stock_quantity=5,
            publication_date=date(1869, 1, 1)
        )

        self.book.authors.add(self.author)
        self.book.categories.add(self.category)

    def test_book_list_page(self):
        response = self.client.get(
            reverse('book_list')
        )

        self.assertEqual(response.status_code, 200)

    def test_book_list_contains_book(self):
        response = self.client.get(
            reverse('book_list')
        )

        self.assertContains(
            response,
            'Война и мир'
        )

    def test_book_detail_page(self):
        response = self.client.get(
            reverse('book_detail', args=[self.book.pk])
        )

        self.assertEqual(response.status_code, 200)

    def test_book_detail_contains_title(self):
        response = self.client.get(
            reverse('book_detail', args=[self.book.pk])
        )

        self.assertContains(
            response,
            'Война и мир'
        )


class ImportBooksFromExcelTest(TestCase):

    def create_excel_file(self):

        workbook = Workbook()
        sheet = workbook.active

        sheet.append([
            "Название",
            "ISBN",
            "Описание",
            "Цена",
            "Количество",
            "Дата публикации",
            "Авторы",
            "Категории"
        ])

        sheet.append([
            "Война и мир",
            "9785170901234",
            "Роман-эпопея",
            1500,
            10,
            date(1869, 1, 1),
            "Лев Толстой",
            "Роман, Классика"
        ])

        file = BytesIO()
        workbook.save(file)
        file.seek(0)

        return file

    def test_import_new_book(self):

        file = self.create_excel_file()

        result = import_books_from_excel(file)

        self.assertEqual(result["books"], 1)
        self.assertEqual(result["authors"], 1)
        self.assertEqual(result["categories"], 2)

        self.assertEqual(Book.objects.count(), 1)
        self.assertEqual(Author.objects.count(), 1)
        self.assertEqual(Category.objects.count(), 2)

        book = Book.objects.get()

        self.assertEqual(
            book.title,
            "Война и мир"
        )

        self.assertEqual(
            book.stock_quantity,
            10
        )

        self.assertEqual(
            book.authors.count(),
            1
        )

        self.assertEqual(
            book.authors.count(),
            1
        )

    def test_update_existing_book(self):

        Book.objects.create(
            title="Война и мир",
            isbn="9785170901234",
            description="Старое описание",
            price=1000,
            stock_quantity=5,
            publication_date=date(1869, 1, 1)
        )

        file = self.create_excel_file()

        result = import_books_from_excel(file)

        self.assertEqual(result["books"], 0)
        self.assertEqual(Book.objects.count(), 1)

        book = Book.objects.get()

        self.assertEqual(
            book.price,
            1500
        )

        self.assertEqual(
            book.stock_quantity,
            10
        )

        self.assertEqual(
            book.description,
            "Роман-эпопея"
        )

        self.assertEqual(
            book.authors.count(),
            1
        )

        self.assertEqual(
            book.categories.count(),
            2
        )

    def test_invalid_headers(self):

        workbook = Workbook()
        sheet = workbook.active

        sheet.append([
            "Название книги",
            "ISBN",
            "Описание",
            "Цена",
            "Количество",
            "Дата публикации",
            "Авторы",
            "Категории"
        ])

        file = BytesIO()
        workbook.save(file)
        file.seek(0)

        with self.assertRaises(ValueError):
            import_books_from_excel(file)

    def test_import_book_with_multiple_authors(self):

        workbook = Workbook()
        sheet = workbook.active

        sheet.append([
            "Название",
            "ISBN",
            "Описание",
            "Цена",
            "Количество",
            "Дата публикации",
            "Авторы",
            "Категории"
        ])

        sheet.append([
            "Пикник на обочине",
            "9785170704324",
            "Научно-фантастическая повесть",
            850,
            20,
            date(1972, 1, 1),
            "Аркадий Стругацкий, Борис Стругацкий",
            "Фантастика"
        ])

        file = BytesIO()
        workbook.save(file)
        file.seek(0)

        result = import_books_from_excel(file)

        self.assertEqual(result["books"], 1)
        self.assertEqual(result["authors"], 2)
        self.assertEqual(result["categories"], 1)

        self.assertEqual(Book.objects.count(), 1)
        self.assertEqual(Author.objects.count(), 2)

        book = Book.objects.get()

        self.assertEqual(
            book.authors.count(),
            2
        )

        self.assertTrue(
            book.authors.filter(
                first_name="Аркадий",
                last_name="Стругацкий"
            ).exists()
        )

        self.assertTrue(
            book.authors.filter(
                first_name="Борис",
                last_name="Стругацкий"
            ).exists()
        )

    def test_import_book_with_multiple_categories(self):

        workbook = Workbook()
        sheet = workbook.active

        sheet.append([
            "Название",
            "ISBN",
            "Описание",
            "Цена",
            "Количество",
            "Дата публикации",
            "Авторы",
            "Категории"
        ])

        sheet.append([
            "Дюна",
            "9785171203530",
            "Научно-фантастический роман",
            2100,
            18,
            date(1965, 8, 1),
            "Фрэнк Герберт",
            "Фантастика, Приключения"
        ])

        file = BytesIO()
        workbook.save(file)
        file.seek(0)

        result = import_books_from_excel(file)

        self.assertEqual(result["books"], 1)
        self.assertEqual(result["authors"], 1)
        self.assertEqual(result["categories"], 2)

        self.assertEqual(Book.objects.count(), 1)
        self.assertEqual(Category.objects.count(), 2)

        book = Book.objects.get()

        self.assertEqual(
            book.categories.count(),
            2
        )

        self.assertTrue(
            book.categories.filter(
                name="Фантастика"
            ).exists()
        )

        self.assertTrue(
            book.categories.filter(
                name="Приключения"
            ).exists()
        )

    def test_import_empty_file(self):

        workbook = Workbook()
        sheet = workbook.active

        sheet.append([
            "Название",
            "ISBN",
            "Описание",
            "Цена",
            "Количество",
            "Дата публикации",
            "Авторы",
            "Категории"
        ])

        file = BytesIO()
        workbook.save(file)
        file.seek(0)

        result = import_books_from_excel(file)

        self.assertEqual(result["books"], 0)
        self.assertEqual(result["authors"], 0)
        self.assertEqual(result["categories"], 0)

        self.assertEqual(Book.objects.count(), 0)
        self.assertEqual(Author.objects.count(), 0)
        self.assertEqual(Category.objects.count(), 0)

    def test_import_ignores_empty_rows(self):

        workbook = Workbook()
        sheet = workbook.active

        sheet.append([
            "Название",
            "ISBN",
            "Описание",
            "Цена",
            "Количество",
            "Дата публикации",
            "Авторы",
            "Категории"
        ])

        sheet.append([
            "Война и мир",
            "9785170901234",
            "Роман-эпопея",
            1500,
            10,
            date(1869, 1, 1),
            "Лев Толстой",
            "Роман"
        ])

        sheet.append([None] * 8)
        sheet.append([None] * 8)
        sheet.append([None] * 8)

        file = BytesIO()
        workbook.save(file)
        file.seek(0)

        result = import_books_from_excel(file)

        self.assertEqual(result["books"], 1)
        self.assertEqual(result["authors"], 1)
        self.assertEqual(result["categories"], 1)

        self.assertEqual(Book.objects.count(), 1)
        self.assertEqual(Author.objects.count(), 1)
        self.assertEqual(Category.objects.count(), 1)


class ExportBooksToExcelTest(TestCase):

    def test_export_empty_catalog(self):

        workbook = export_books_to_excel()

        sheet = workbook.active

        headers = [
            cell.value
            for cell in sheet[1]
        ]

        self.assertEqual(sheet.max_row, 1)
        self.assertEqual(sheet.max_column, 8)

        self.assertEqual(
            headers,
            [
                "Название",
                "ISBN",
                "Описание",
                "Цена",
                "Количество",
                "Дата публикации",
                "Авторы",
                "Категории"
            ]
        )

    def test_export_single_book(self):

        author = Author.objects.create(
            first_name="Лев",
            last_name="Толстой"
        )

        category = Category.objects.create(
            name="Классика"
        )

        book = Book.objects.create(
            title="Война и мир",
            isbn="9785170901234",
            description="Роман",
            price=1500,
            stock_quantity=10,
            publication_date=date(1869, 1, 1)
        )

        book.authors.add(author)
        book.categories.add(category)

        workbook = export_books_to_excel()

        sheet = workbook.active

        self.assertEqual(sheet.max_row, 2)

        self.assertEqual(sheet["A2"].value, "Война и мир")
        self.assertEqual(sheet["B2"].value, "9785170901234")
        self.assertEqual(sheet["C2"].value, "Роман")
        self.assertEqual(sheet["D2"].value, 1500)
        self.assertEqual(sheet["E2"].value, 10)
        self.assertEqual(sheet["F2"].value, "1869-01-01")
        self.assertEqual(sheet["G2"].value, "Лев Толстой")
        self.assertEqual(sheet["H2"].value, "Классика")

    def test_export_multiple_authors(self):

        first = Author.objects.create(
            first_name="Аркадий",
            last_name="Стругацкий"
        )

        second = Author.objects.create(
            first_name="Борис",
            last_name="Стругацкий"
        )

        book = Book.objects.create(
            title="Пикник на обочине",
            isbn="9785170704324",
            description="Фантастика",
            price=850,
            stock_quantity=20,
            publication_date=date(1972, 1, 1)
        )

        book.authors.add(first, second)

        workbook = export_books_to_excel()

        authors = workbook.active["G2"].value

        self.assertIn("Аркадий Стругацкий", authors)
        self.assertIn("Борис Стругацкий", authors)

    def test_export_multiple_categories(self):

        author = Author.objects.create(
            first_name="Фрэнк",
            last_name="Герберт",
        )

        first = Category.objects.create(name="Фантастика")
        second = Category.objects.create(name="Приключения")

        book = Book.objects.create(
            title="Дюна",
            isbn="9785171203530",
            description="Роман",
            price=2100,
            stock_quantity=18,
            publication_date=date(1965, 8, 1),
        )

        book.authors.add(author)
        book.categories.add(first, second)

        workbook = export_books_to_excel()

        categories = workbook.active["H2"].value

        self.assertIn("Фантастика", categories)
        self.assertIn("Приключения", categories)
