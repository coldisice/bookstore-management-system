from datetime import datetime
from django.db import transaction

from openpyxl import load_workbook, Workbook

from .models import Author, Book, Category

REQUIRED_COLUMNS = [
    "Название",
    "ISBN",
    "Описание",
    "Цена",
    "Количество",
    "Дата публикации",
    "Авторы",
    "Категории"
]


@transaction.atomic
def import_books_from_excel(file):

    workbook = load_workbook(file)
    sheet = workbook.active

    headers = [
        cell.value
        for cell in sheet[1]
    ]

    if headers != REQUIRED_COLUMNS:
        raise ValueError(
            "Неверная структура Excel-файла."
        )

    books_created = 0
    authors_created = 0
    categories_created = 0

    for row in sheet.iter_rows(
        min_row=2,
        values_only=True
    ):

        if all(value is None for value in row):
            continue

        (
            title,
            isbn,
            description,
            price,
            stock_quantity,
            publication_date,
            authors,
            categories
        ) = row

        if not title:
            raise ValueError("Не заполнено поле 'Название'.")

        if not isbn:
            raise ValueError(
                f"У книги '{title} отсутствует ISBN."
            )

        if isinstance(publication_date, datetime):
            publication_date = publication_date.date()

        book, created = Book.objects.update_or_create(
            isbn=isbn,
            defaults={
                "title": title,
                "description": description,
                "price": price,
                "stock_quantity": stock_quantity,
                "publication_date": publication_date,
            },
        )

        if created:
            books_created += 1

        book.authors.clear()

        for author_name in authors.split(","):

            author_name = author_name.strip()

            if not author_name:
                continue

            parts = author_name.split()

            if len(parts) == 1:
                first_name = parts[0]
                last_name = ""
            else:
                first_name = " ".join(parts[:-1])
                last_name = parts[-1]

            author, created = Author.objects.get_or_create(
                first_name=first_name,
                last_name=last_name,
            )

            if created:
                authors_created += 1

            book.authors.add(author)

        book.categories.clear()

        for category_name in categories.split(","):

            category_name = category_name.strip()

            if not category_name:
                continue

            category, created = Category.objects.get_or_create(
                name=category_name
            )

            if created:
                categories_created += 1

            book.categories.add(category)

    return {
        "books": books_created,
        "authors": authors_created,
        "categories": categories_created,
    }


def export_books_to_excel():

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Книги"

    sheet.append([
        "Название",
        "ISBN",
        "Описание",
        "Цена",
        "Количество",
        "Дата публикации",
        "Авторы",
        "Категории"
    ])

    books = Book.objects.prefetch_related(
        "authors",
        "categories",
    )

    for book in books:

        authors = ", ".join(
            str(author)
            for author in book.authors.all()
        )

        categories = ", ".join(
            category.name
            for category in book.categories.all()
        )

        publication_date = book.publication_date

        if publication_date.year < 1900:
            publication_date = str(publication_date)

        sheet.append([
            book.title,
            book.isbn,
            book.description,
            float(book.price),
            book.stock_quantity,
            publication_date,
            authors,
            categories
        ])

    return workbook
